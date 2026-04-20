from __future__ import annotations

import io

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from core.auth import get_current_user, require_admin, require_faculty_or_admin
from core.database import get_db
from models.tables import (
    Assignment,
    AssignmentType,
    CodingAssessment,
    CodingSubmission,
    Course,
    Enrollment,
    QuizAttempt,
    Submission,
    User,
    UserRole,
)
from schemas.requests import (
    CourseCreate,
    CourseOut,
    CourseUpdate,
    EnrollRequest,
    MessageResponse,
    UserBrief,
)

router = APIRouter(prefix="/courses", tags=["courses"])


async def _get_course_or_404(db: AsyncSession, course_id: int) -> Course:
    result = await db.execute(
        select(Course)
        .options(selectinload(Course.faculty))
        .where(Course.id == course_id)
    )
    course = result.scalar_one_or_none()
    if course is None:
        raise HTTPException(status_code=404, detail="Course not found")
    return course


def _serialize(
    course: Course,
    *,
    enrolled: bool | None = None,
    enrollment_count: int | None = None,
) -> CourseOut:
    return CourseOut(
        id=course.id,
        title=course.title,
        code=course.code,
        description=course.description,
        department_id=course.department_id,
        faculty_id=course.faculty_id,
        faculty_name=course.faculty.name if course.faculty else None,
        semester=course.semester,
        created_at=course.created_at,
        enrolled=enrolled,
        enrollment_count=enrollment_count,
    )


@router.get("", response_model=list[CourseOut])
async def list_courses(
    mine: bool = False,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> list[CourseOut]:
    stmt = select(Course).options(selectinload(Course.faculty)).order_by(Course.created_at.desc())

    if user.role == UserRole.student:
        if mine:
            stmt = stmt.join(Enrollment, Enrollment.course_id == Course.id).where(
                Enrollment.student_id == user.id
            )
        # Students otherwise see all courses so they can self-enroll.
    elif user.role == UserRole.faculty and mine:
        stmt = stmt.where(Course.faculty_id == user.id)

    result = await db.execute(stmt)
    courses = result.scalars().unique().all()

    # Enrollment info for students
    enrolled_ids: set[int] = set()
    if user.role == UserRole.student:
        er = await db.execute(
            select(Enrollment.course_id).where(Enrollment.student_id == user.id)
        )
        enrolled_ids = {row[0] for row in er.all()}

    # Enrollment counts (admin + faculty)
    counts: dict[int, int] = {}
    if user.role in (UserRole.admin, UserRole.faculty):
        ids = [c.id for c in courses]
        if ids:
            cr = await db.execute(
                select(Enrollment.course_id, func.count(Enrollment.id))
                .where(Enrollment.course_id.in_(ids))
                .group_by(Enrollment.course_id)
            )
            counts = {cid: n for cid, n in cr.all()}

    return [
        _serialize(
            c,
            enrolled=(c.id in enrolled_ids) if user.role == UserRole.student else None,
            enrollment_count=counts.get(c.id),
        )
        for c in courses
    ]


@router.get("/{course_id}", response_model=CourseOut)
async def get_course(
    course_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> CourseOut:
    course = await _get_course_or_404(db, course_id)
    enrolled: bool | None = None
    enrollment_count: int | None = None
    if user.role == UserRole.student:
        er = await db.execute(
            select(Enrollment.id).where(
                Enrollment.course_id == course_id, Enrollment.student_id == user.id
            )
        )
        enrolled = er.scalar_one_or_none() is not None
    else:
        cr = await db.execute(
            select(func.count(Enrollment.id)).where(Enrollment.course_id == course_id)
        )
        enrollment_count = cr.scalar_one()
    return _serialize(course, enrolled=enrolled, enrollment_count=enrollment_count)


@router.post("", response_model=CourseOut, status_code=status.HTTP_201_CREATED)
async def create_course(
    payload: CourseCreate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_faculty_or_admin),
) -> CourseOut:
    faculty_id = payload.faculty_id or (user.id if user.role == UserRole.faculty else None)
    if faculty_id is not None:
        fr = await db.execute(
            select(User).where(User.id == faculty_id, User.role == UserRole.faculty)
        )
        if fr.scalar_one_or_none() is None:
            raise HTTPException(status_code=400, detail="faculty_id must reference a faculty user")

    course = Course(
        title=payload.title,
        code=payload.code.upper(),
        description=payload.description,
        department_id=payload.department_id,
        faculty_id=faculty_id,
        semester=payload.semester,
    )
    db.add(course)
    try:
        await db.commit()
    except IntegrityError:
        await db.rollback()
        raise HTTPException(status_code=409, detail="Course code already exists")
    await db.refresh(course, attribute_names=["faculty"])
    return _serialize(course, enrollment_count=0)


@router.patch("/{course_id}", response_model=CourseOut)
async def update_course(
    course_id: int,
    payload: CourseUpdate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_faculty_or_admin),
) -> CourseOut:
    course = await _get_course_or_404(db, course_id)

    # Faculty may only edit their own course; and cannot reassign faculty.
    if user.role == UserRole.faculty:
        if course.faculty_id != user.id:
            raise HTTPException(status_code=403, detail="Not your course")
        if payload.faculty_id is not None and payload.faculty_id != user.id:
            raise HTTPException(status_code=403, detail="Faculty cannot reassign courses")

    data = payload.model_dump(exclude_unset=True)
    if "faculty_id" in data and data["faculty_id"] is not None:
        fr = await db.execute(
            select(User).where(User.id == data["faculty_id"], User.role == UserRole.faculty)
        )
        if fr.scalar_one_or_none() is None:
            raise HTTPException(status_code=400, detail="faculty_id must reference a faculty user")
    for k, v in data.items():
        setattr(course, k, v)

    await db.commit()
    await db.refresh(course, attribute_names=["faculty"])
    return _serialize(course)


@router.delete("/{course_id}", response_model=MessageResponse)
async def delete_course(
    course_id: int,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_admin),
) -> MessageResponse:
    course = await _get_course_or_404(db, course_id)
    await db.delete(course)
    await db.commit()
    return MessageResponse(detail="Course deleted")


# ---------- Enrollments ----------


@router.post("/{course_id}/enroll", response_model=MessageResponse)
async def enroll(
    course_id: int,
    payload: EnrollRequest | None = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> MessageResponse:
    await _get_course_or_404(db, course_id)

    target_id: int
    if user.role == UserRole.student:
        if payload and payload.student_id and payload.student_id != user.id:
            raise HTTPException(status_code=403, detail="Students can only enroll themselves")
        target_id = user.id
    else:
        if not payload or not payload.student_id:
            raise HTTPException(status_code=400, detail="student_id is required")
        sr = await db.execute(
            select(User).where(User.id == payload.student_id, User.role == UserRole.student)
        )
        if sr.scalar_one_or_none() is None:
            raise HTTPException(status_code=400, detail="student_id must reference a student")
        target_id = payload.student_id

    enrollment = Enrollment(student_id=target_id, course_id=course_id)
    db.add(enrollment)
    try:
        await db.commit()
    except IntegrityError:
        await db.rollback()
        raise HTTPException(status_code=409, detail="Already enrolled")
    return MessageResponse(detail="Enrolled")


@router.delete("/{course_id}/enroll", response_model=MessageResponse)
async def unenroll(
    course_id: int,
    student_id: int | None = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> MessageResponse:
    target_id = user.id if user.role == UserRole.student else student_id
    if target_id is None:
        raise HTTPException(status_code=400, detail="student_id is required")

    if user.role == UserRole.student and student_id and student_id != user.id:
        raise HTTPException(status_code=403, detail="Students can only unenroll themselves")

    result = await db.execute(
        select(Enrollment).where(
            Enrollment.course_id == course_id, Enrollment.student_id == target_id
        )
    )
    enrollment = result.scalar_one_or_none()
    if enrollment is None:
        raise HTTPException(status_code=404, detail="Not enrolled")
    await db.delete(enrollment)
    await db.commit()
    return MessageResponse(detail="Unenrolled")


@router.get("/{course_id}/students", response_model=list[UserBrief])
async def list_course_students(
    course_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_faculty_or_admin),
) -> list[UserBrief]:
    course = await _get_course_or_404(db, course_id)
    if user.role == UserRole.faculty and course.faculty_id != user.id:
        raise HTTPException(status_code=403, detail="Not your course")

    result = await db.execute(
        select(User)
        .join(Enrollment, Enrollment.student_id == User.id)
        .where(Enrollment.course_id == course_id)
        .order_by(User.name)
    )
    return [UserBrief.model_validate(u) for u in result.scalars().all()]


# ---------- Performance export (faculty/admin) ----------


@router.get("/{course_id}/performance.xlsx")
async def export_performance(
    course_id: int,
    student_ids: str | None = Query(
        default=None,
        description="Comma-separated student ids to export (default: all enrolled)",
    ),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_faculty_or_admin),
) -> StreamingResponse:
    """Excel export of student performance for a course.

    Sheets:
    - Summary — one row per student with totals across all assessment types
    - Assignments — per (student, assignment) with marks
    - Quizzes — per (student, quiz) with score
    - Coding — per (student, coding assessment) with best score
    """
    # Lazy import keeps cold-start light if export is unused.
    from openpyxl import Workbook

    course = await _get_course_or_404(db, course_id)
    if user.role == UserRole.faculty and course.faculty_id != user.id:
        raise HTTPException(status_code=403, detail="Not your course")

    # Roster
    roster_res = await db.execute(
        select(User)
        .join(Enrollment, Enrollment.student_id == User.id)
        .where(Enrollment.course_id == course_id, User.role == UserRole.student)
        .order_by(User.name)
    )
    roster: list[User] = list(roster_res.scalars().all())

    if student_ids:
        try:
            ids = {int(s) for s in student_ids.split(",") if s.strip()}
        except ValueError:
            raise HTTPException(
                status_code=400,
                detail="student_ids must be comma-separated integers",
            )
        roster = [s for s in roster if s.id in ids]

    if not roster:
        raise HTTPException(
            status_code=404, detail="No students match this filter"
        )

    student_id_set = {s.id for s in roster}

    # Course assignments (file + quiz)
    asg_res = await db.execute(
        select(Assignment)
        .where(Assignment.course_id == course_id)
        .order_by(Assignment.due_date.asc())
    )
    asgs: list[Assignment] = list(asg_res.scalars().all())
    file_asgs = [a for a in asgs if a.type == AssignmentType.file]
    quiz_asgs = [a for a in asgs if a.type == AssignmentType.quiz]

    # Submissions (file)
    sub_res = await db.execute(
        select(Submission).where(
            Submission.assignment_id.in_([a.id for a in file_asgs] or [0]),
            Submission.student_id.in_(student_id_set),
        )
    )
    sub_by_pair: dict[tuple[int, int], Submission] = {
        (s.assignment_id, s.student_id): s for s in sub_res.scalars().all()
    }

    # Quiz attempts
    qa_res = await db.execute(
        select(QuizAttempt).where(
            QuizAttempt.assignment_id.in_([a.id for a in quiz_asgs] or [0]),
            QuizAttempt.student_id.in_(student_id_set),
        )
    )
    qa_by_pair: dict[tuple[int, int], QuizAttempt] = {
        (a.assignment_id, a.student_id): a for a in qa_res.scalars().all()
    }

    # Coding assessments (graded for this course)
    coding_res = await db.execute(
        select(CodingAssessment)
        .where(
            CodingAssessment.course_id == course_id,
            CodingAssessment.is_practice.is_(False),
        )
        .order_by(CodingAssessment.created_at.asc())
    )
    coding_asgs: list[CodingAssessment] = list(coding_res.scalars().all())

    # Best coding score per (student, assessment)
    csub_res = await db.execute(
        select(
            CodingSubmission.assessment_id,
            CodingSubmission.student_id,
            func.max(CodingSubmission.score),
        )
        .where(
            CodingSubmission.assessment_id.in_(
                [c.id for c in coding_asgs] or [0]
            ),
            CodingSubmission.student_id.in_(student_id_set),
        )
        .group_by(CodingSubmission.assessment_id, CodingSubmission.student_id)
    )
    csub_by_pair: dict[tuple[int, int], int] = {
        (aid, sid): int(best or 0) for aid, sid, best in csub_res.all()
    }

    # Build workbook
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append([
        "Student ID",
        "Name",
        "Email",
        "File submitted / total",
        "File graded avg %",
        "Quiz submitted / total",
        "Quiz avg %",
        "Coding attempted / total",
        "Coding avg %",
    ])

    for s in roster:
        f_total = len(file_asgs)
        f_sub = sum(
            1 for a in file_asgs if (a.id, s.id) in sub_by_pair
        )
        f_graded_pcts: list[float] = []
        for a in file_asgs:
            sub = sub_by_pair.get((a.id, s.id))
            if sub and sub.marks is not None and a.max_marks > 0:
                f_graded_pcts.append(100 * sub.marks / a.max_marks)
        f_avg = (
            round(sum(f_graded_pcts) / len(f_graded_pcts), 1)
            if f_graded_pcts
            else None
        )

        q_total = len(quiz_asgs)
        q_sub = sum(
            1
            for a in quiz_asgs
            if (a.id, s.id) in qa_by_pair
            and qa_by_pair[(a.id, s.id)].submitted_at is not None
        )
        q_pcts: list[float] = []
        for a in quiz_asgs:
            at = qa_by_pair.get((a.id, s.id))
            if at and at.score is not None and at.max_score:
                q_pcts.append(100 * at.score / at.max_score)
        q_avg = round(sum(q_pcts) / len(q_pcts), 1) if q_pcts else None

        c_total = len(coding_asgs)
        c_attempted = sum(
            1 for c in coding_asgs if (c.id, s.id) in csub_by_pair
        )
        c_pcts: list[float] = []
        for c in coding_asgs:
            best = csub_by_pair.get((c.id, s.id))
            if best is not None and c.max_score > 0:
                c_pcts.append(100 * best / c.max_score)
        c_avg = round(sum(c_pcts) / len(c_pcts), 1) if c_pcts else None

        summary.append([
            s.id,
            s.name,
            s.email,
            f"{f_sub}/{f_total}",
            f_avg if f_avg is not None else "",
            f"{q_sub}/{q_total}",
            q_avg if q_avg is not None else "",
            f"{c_attempted}/{c_total}",
            c_avg if c_avg is not None else "",
        ])

    # Assignments sheet
    asg_sheet = wb.create_sheet("Assignments")
    asg_sheet.append([
        "Student ID",
        "Student",
        "Assignment",
        "Due",
        "Max",
        "Marks",
        "Submitted at",
        "Graded at",
        "Feedback",
    ])
    for s in roster:
        for a in file_asgs:
            sub = sub_by_pair.get((a.id, s.id))
            asg_sheet.append([
                s.id,
                s.name,
                a.title,
                a.due_date.isoformat() if a.due_date else "",
                a.max_marks,
                sub.marks if sub and sub.marks is not None else "",
                sub.submitted_at.isoformat() if sub and sub.submitted_at else "",
                sub.graded_at.isoformat() if sub and sub.graded_at else "",
                (sub.feedback or "") if sub else "",
            ])

    # Quizzes sheet
    quiz_sheet = wb.create_sheet("Quizzes")
    quiz_sheet.append([
        "Student ID",
        "Student",
        "Quiz",
        "Due",
        "Score",
        "Max score",
        "Started at",
        "Submitted at",
    ])
    for s in roster:
        for a in quiz_asgs:
            at = qa_by_pair.get((a.id, s.id))
            quiz_sheet.append([
                s.id,
                s.name,
                a.title,
                a.due_date.isoformat() if a.due_date else "",
                at.score if at and at.score is not None else "",
                at.max_score if at and at.max_score is not None else "",
                at.started_at.isoformat() if at and at.started_at else "",
                at.submitted_at.isoformat() if at and at.submitted_at else "",
            ])

    # Coding sheet
    coding_sheet = wb.create_sheet("Coding")
    coding_sheet.append([
        "Student ID",
        "Student",
        "Assessment",
        "Due",
        "Max",
        "Best score",
    ])
    for s in roster:
        for c in coding_asgs:
            best = csub_by_pair.get((c.id, s.id))
            coding_sheet.append([
                s.id,
                s.name,
                c.title,
                c.due_date.isoformat() if c.due_date else "",
                c.max_score,
                best if best is not None else "",
            ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"performance_course_{course_id}.xlsx"
    return StreamingResponse(
        buf,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
